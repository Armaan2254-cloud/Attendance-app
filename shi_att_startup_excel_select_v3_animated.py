from kivy.app import App

from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.uix.behaviors import ButtonBehavior
from kivy.clock import Clock
from kivy.animation import Animation

from kivy.graphics import Color, RoundedRectangle, Line

from openpyxl import load_workbook

from datetime import datetime, date, timedelta

import os


# ============================================================
# OPTIONAL ANDROID STORAGE PERMISSION
# ============================================================
# Imported only when available, so desktop/testing startup is safe.
try:
    from android.permissions import request_permissions, Permission
except Exception:
    request_permissions = None
    Permission = None


# ============================================================
# EXCEL FILE
# ============================================================

# No Excel file is hard-coded.
# The user selects the workbook from Download when the app starts.
# The selected path is stored in self.file_path.



# ============================================================
# STATUS BUTTON
# ============================================================

class StatusButton(ButtonBehavior, Label):

    def __init__(
        self,
        status_code="P",
        selected_color=(0.45, 0.90, 0.45, 1),
        border_color=(0.45, 0.90, 0.45, 1),
        **kwargs
    ):

        super().__init__(**kwargs)

        self.status_code = status_code

        self.selected_color = selected_color
        self.border_color = border_color

        self.font_size = 18
        self.bold = True

        # Text color
        self.color = border_color

        # ----------------------------------------------------
        # Background
        # ----------------------------------------------------

        with self.canvas.before:

            # Background
            self.bg_color = Color(
                rgba=(0, 0, 0, 0)
            )

            self.bg = RoundedRectangle(
                pos=self.pos,
                size=self.size,
                radius=[8]
            )

            # Border
            self.line_color = Color(
                rgba=border_color
            )

            self.border = Line(
                rounded_rectangle=(
                    self.x,
                    self.y,
                    self.width,
                    self.height,
                    8
                ),
                width=1.5
            )

        self.bind(
            pos=self.update_graphics,
            size=self.update_graphics
        )

        # Small press animation for P/A/H buttons.
        self.bind(
            on_press=self.animate_press,
            on_release=self.animate_release
        )

    def animate_press(self, *args):
        Animation.cancel_all(self, "font_size")
        Animation(font_size=15, duration=0.06).start(self)

    def animate_release(self, *args):
        Animation(font_size=18, duration=0.10).start(self)

    # ========================================================
    # GRAPHICS UPDATE
    # ========================================================

    def update_graphics(self, *args):

        self.bg.pos = self.pos
        self.bg.size = self.size

        self.border.rounded_rectangle = (
            self.x,
            self.y,
            self.width,
            self.height,
            8
        )

    # ========================================================
    # SELECT / DESELECT
    # ========================================================

    def set_selected(self, selected):

        if selected:

            # Filled background
            self.bg_color.rgba = self.selected_color

            # Dark text
            self.color = (0, 0, 0, 1)

        else:

            # Transparent background
            self.bg_color.rgba = (
                0,
                0,
                0,
                0
            )

            # Colored text
            self.color = self.border_color


# ============================================================
# MAIN APP
# ============================================================

class AttendanceApp(App):

    def build(self):

        self.records = []
        self.file_path = None
        self.startup_file_popup = None

        self.current_date = datetime.now().date()

        # ====================================================
        # MAIN LAYOUT
        # ====================================================

        main = BoxLayout(
            orientation="vertical",
            padding=12,
            spacing=6
        )

        # ====================================================
        # TITLE
        # ====================================================

        title = Label(
            text="ATTENDANCE MANAGER",
            font_size=50,
            size_hint_y=None,
            height=55
        )

        main.add_widget(title)

        # Initial entrance animation.
        title.opacity = 0
        Animation(opacity=1, duration=0.35).start(title)

        # ====================================================
        # CHANGE EXCEL FILE
        # ====================================================

        change_file = Button(
            text="CHANGE EXCEL FILE",
            font_size=24,
            size_hint_y=None,
            height=55
        )
        change_file.bind(on_press=self.change_excel_file)
        main.add_widget(change_file)
        change_file.opacity = 0
        Clock.schedule_once(lambda dt: Animation(opacity=1, duration=0.25).start(change_file), 0.08)

        # ====================================================
        # DATE
        # ====================================================

        self.date_label = Label(
            text="",
            font_size=31,
            size_hint_y=None,
            height=42
        )

        main.add_widget(
            self.date_label
        )

        # ====================================================
        # SUMMARY
        # ====================================================

        summary_box = BoxLayout(
            size_hint_y=None,
            height=52,
            spacing=12
        )

        self.present_label = Label(
            text="Present: 0",
            font_size=28,
            color=(0.35, 1, 0.35, 1)
        )

        self.absent_label = Label(
            text="Absent: 0",
            font_size=28,
            color=(1, 0.35, 0.35, 1)
        )

        self.holiday_label = Label(
            text="Holiday: 0",
            font_size=28,
            color=(0.40, 0.60, 1, 1)
        )

        summary_box.add_widget(
            self.present_label
        )

        summary_box.add_widget(
            self.absent_label
        )

        summary_box.add_widget(
            self.holiday_label
        )

        main.add_widget(
            summary_box
        )

        # ====================================================
        # DATE NAVIGATION
        # ====================================================

        nav = BoxLayout(
            size_hint_y=None,
            height=65,
            spacing=10
        )

        previous = Button(
            text="< PREVIOUS",
            font_size=30
        )

        today = Button(
            text="TODAY",
            font_size=30
        )

        next_date = Button(
            text="NEXT >",
            font_size=30
        )

        previous.bind(
            on_press=self.previous_date
        )

        today.bind(
            on_press=self.go_today
        )

        next_date.bind(
            on_press=self.next_date
        )

        nav.add_widget(previous)
        nav.add_widget(today)
        nav.add_widget(next_date)

        main.add_widget(nav)
        nav.opacity = 0
        Clock.schedule_once(lambda dt: Animation(opacity=1, duration=0.25).start(nav), 0.15)

        # ====================================================
        # STUDENT SCROLL
        # ====================================================

        self.scroll = ScrollView(
            do_scroll_x=False
        )

        self.student_list = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            spacing=8,
            padding=2
        )

        self.student_list.bind(
            minimum_height=self.student_list.setter(
                "height"
            )
        )

        self.scroll.add_widget(
            self.student_list
        )

        main.add_widget(
            self.scroll
        )

        # ====================================================
        # SAVE BUTTON
        # ====================================================

        save = Button(
            text="SAVE ATTENDANCE",
            font_size=41,
            size_hint_y=None,
            height=65
        )

        save.bind(
            on_press=self.save_current
        )

        main.add_widget(save)
        save.opacity = 0
        Clock.schedule_once(lambda dt: Animation(opacity=1, duration=0.25).start(save), 0.20)

        # ====================================================
        # REPORT BUTTON
        # ====================================================

        report = Button(
            text="ATTENDANCE REPORT",
            font_size=41,
            size_hint_y=None,
            height=65
        )

        report.bind(
            on_press=self.show_report
        )

        main.add_widget(report)
        report.opacity = 0
        Clock.schedule_once(lambda dt: Animation(opacity=1, duration=0.25).start(report), 0.25)

        # ====================================================
        # STARTUP FILE SELECTION
        # ====================================================
        # Do NOT load any hard-coded Excel file.
        # First let Kivy finish creating the UI, then open the
        # Excel selector. This avoids startup crashes caused by
        # opening the file chooser during build().

        self.date_label.text = (
            self.current_date.strftime("%d %B %Y")
            + " - SELECT EXCEL FILE"
        )

        Clock.schedule_once(
            lambda dt: self.open_excel_selector(startup=True),
            0.30
        )

        return main

    # ========================================================
    # STORAGE / EXCEL FILE HELPERS
    # ========================================================

    def request_storage_permission_safe(self):
        # Never let a permission import/request crash app startup.
        if request_permissions is None or Permission is None:
            return

        try:
            permissions = []
            if hasattr(Permission, "READ_EXTERNAL_STORAGE"):
                permissions.append(Permission.READ_EXTERNAL_STORAGE)
            if hasattr(Permission, "WRITE_EXTERNAL_STORAGE"):
                permissions.append(Permission.WRITE_EXTERNAL_STORAGE)
            if permissions:
                request_permissions(permissions)
        except Exception:
            pass

    def get_download_xlsx_files(self):
        folder = "/storage/emulated/0/Excel"

        if not os.path.isdir(folder):
            raise Exception("Download folder access nahi mil rahi.")

        return sorted(
            [
                f for f in os.listdir(folder)
                if f.lower().endswith(".xlsx")
                and os.path.isfile(os.path.join(folder, f))
            ],
            key=str.lower
        )

    # ========================================================
    # CHANGE / SELECT EXCEL FILE
    # ========================================================

    def change_excel_file(self, instance):
        self.open_excel_selector(startup=False)

    def open_excel_selector(self, startup=False):
        """Show all .xlsx files directly inside Download.

        startup=True is used only for the first automatic selection popup.
        No workbook is loaded until the user taps a file.
        """

        try:
            files = self.get_download_xlsx_files()
        except Exception as e:
            self.show_error(str(e))
            return

        if not files:
            self.show_error(
                "Excel folder mein koi .xlsx file nahi mili."
            )
            return

        box = BoxLayout(
            orientation="vertical",
            spacing=5,
            padding=8
        )

        scroll = ScrollView(
            do_scroll_x=False
        )

        file_list = BoxLayout(
            orientation="vertical",
            size_hint_y=None,
            spacing=5
        )

        file_list.bind(
            minimum_height=file_list.setter("height")
        )

        scroll.add_widget(file_list)
        box.add_widget(scroll)

        # During startup, don't provide a Cancel button: the first action
        # of the app is to choose the workbook. After startup, the normal
        # CHANGE EXCEL FILE button simply opens the same selector.
        if not startup:
            close = Button(
                text="CANCEL",
                font_size=20,
                size_hint_y=None,
                height=50
            )
            box.add_widget(close)

        popup = Popup(
            title="Select Excel File",
            content=box,
            size_hint=(0.92, 0.85),
            auto_dismiss=not startup
        )

        self.startup_file_popup = popup if startup else None

        if not startup:
            close.bind(on_press=popup.dismiss)

        folder = "/storage/emulated/0/Excel"

        for filename in files:

            button = Button(
                text=filename,
                font_size=20,
                size_hint_y=None,
                height=55
            )

            def choose(instance, selected=filename):
                selected_path = os.path.join(folder, selected)

                # Validate BEFORE changing the active file.
                try:
                    self.validate_excel_file(selected_path)
                except Exception as e:
                    self.show_error(
                        "Excel file load nahi ho sakti:\n\n" + str(e)
                    )
                    return

                # Only now make this workbook the active workbook.
                self.file_path = selected_path

                if popup == self.startup_file_popup:
                    self.startup_file_popup = None

                popup.dismiss()
                self.load_date(self.current_date)

            button.bind(on_press=choose)
            file_list.add_widget(button)

        popup.open()
        box.opacity = 0
        Clock.schedule_once(
            lambda dt: Animation(opacity=1, duration=0.25).start(box),
            0
        )

    def validate_excel_file(self, path):
        if not os.path.isfile(path):
            raise Exception("Excel file not found.")

        wb = load_workbook(
            path,
            read_only=True,
            data_only=False
        )

        try:
            if "Sheet2" not in wb.sheetnames:
                raise Exception(
                    "Is Excel file mein 'Sheet2' sheet nahi hai."
                )

            ws = wb["Sheet2"]
            month_value = ws["B2"].value

            if month_value is None or str(month_value).strip() == "":
                raise Exception(
                    "Sheet2 ke B2 cell mein month nahi mila."
                )
        finally:
            wb.close()

    # ========================================================
    # CHECK FILE
    # ========================================================

    def check_file(self):

        if not self.file_path:
            raise Exception(
                "Pehle Download folder se Excel file select karein."
            )

        if not os.path.isfile(self.file_path):
            raise Exception(
                "Excel file not found:\n\n"
                + self.file_path
                + "\n\nCHANGE EXCEL FILE dabakar Download folder se .xlsx file select karein."
            )

    # ========================================================
    # GET EXCEL MONTH
    # ========================================================

    def get_excel_month_year(self):

        self.check_file()

        wb = load_workbook(
            self.file_path,
            data_only=False
        )

        ws = wb["Sheet2"]

        month_value = ws["B2"].value

        wb.close()

        if month_value is None:

            raise Exception(
                "Excel ke B2 cell mein month nahi mila."
            )

        month_text = str(
            month_value
        ).strip()

        month_number = None

        # Full month
        try:

            month_number = datetime.strptime(
                month_text,
                "%B"
            ).month

        except ValueError:

            pass

        # Short month
        if month_number is None:

            try:

                month_number = datetime.strptime(
                    month_text,
                    "%b"
                ).month

            except ValueError:

                pass

        # Numeric month
        if month_number is None:

            try:

                month_number = int(
                    month_text
                )

            except:

                pass

        if month_number is None:

            raise Exception(
                "B2 mein valid month nahi mila:\n"
                + month_text
            )

        # ----------------------------------------------------
        # Current Excel year
        # ----------------------------------------------------

        year = 2026

        return month_number, year

    # ========================================================
    # FIND DATE COLUMN
    # ========================================================

    def find_date_column(self, target_date):

        month_number, year = (
            self.get_excel_month_year()
        )

        if target_date.year != year:
            return None

        if target_date.month != month_number:
            return None

        first_date = date(
            year,
            month_number,
            1
        )

        difference = (
            target_date - first_date
        ).days

        if difference < 0:
            return None

        # D = 1st date
        first_column = 4

        column = (
            first_column
            + difference
        )

        wb = load_workbook(
            self.file_path,
            data_only=False
        )

        ws = wb["Sheet2"]

        if column > ws.max_column:

            wb.close()

            return None

        wb.close()

        return column

    # ========================================================
    # LOAD DATE
    # ========================================================

    def load_date(self, selected_date):

        # Smoothly fade the student area while switching dates.
        if hasattr(self, "student_list"):
            Animation.cancel_all(self.student_list, "opacity")
            Animation(opacity=0.15, duration=0.08).start(self.student_list)

        self.current_date = selected_date

        self.records = []

        self.student_list.clear_widgets()

        self.date_label.text = (
            selected_date.strftime(
                "%d %B %Y"
            )
        )

        try:

            column = self.find_date_column(
                selected_date
            )

            if column is None:

                self.date_label.text = (
                    selected_date.strftime(
                        "%d %B %Y"
                    )
                    + " - DATE NOT FOUND"
                )

                self.present_label.text = (
                    "Present: 0"
                )

                self.absent_label.text = (
                    "Absent: 0"
                )

                self.holiday_label.text = (
                    "Holiday: 0"
                )

                return

            wb = load_workbook(
                self.file_path,
                data_only=False
            )

            ws = wb["Sheet2"]

            # =================================================
            # STUDENTS
            # =================================================

            for row in range(
                5,
                ws.max_row + 1
            ):

                name = ws.cell(
                    row,
                    2
                ).value

                if name is None:
                    continue

                name = str(
                    name
                ).strip()

                if name == "":
                    continue

                # =============================================
                # EXISTING VALUE
                # =============================================

                existing = ws.cell(
                    row,
                    column
                ).value

                if existing is None:
                    existing = ""

                existing = str(
                    existing
                ).strip().upper()

                if existing not in [
                    "P",
                    "A",
                    "H"
                ]:

                    existing = ""

                # =============================================
                # STUDENT ROW
                # =============================================

                row_box = BoxLayout(
                    size_hint_y=None,
                    height=78,
                    spacing=10,
                    padding=10
                )

                # =============================================
                # NAME
                # =============================================

                name_label = Label(
                    text=name,
                    font_size=31,
                    halign="left",
                    valign="middle",
                    size_hint_x=0.40
                )

                name_label.bind(
                    size=lambda obj, value:
                    setattr(
                        obj,
                        "text_size",
                        value
                    )
                )

                # =============================================
                # STATUS LABEL
                # =============================================

                status = Label(
                    text=existing,
                    font_size=34,
                    bold=True,
                    size_hint_x=0.10
                )

                if existing == "P":

                    status.color = (
                        0.35,
                        1,
                        0.35,
                        1
                    )

                elif existing == "A":

                    status.color = (
                        1,
                        0.35,
                        0.35,
                        1
                    )

                elif existing == "H":

                    status.color = (
                        0.40,
                        0.60,
                        1,
                        1
                    )

                # =============================================
                # P BUTTON
                # =============================================

                p = StatusButton(
                    text="P",
                    status_code="P",

                    selected_color=(
                        0.45,
                        0.90,
                        0.45,
                        1
                    ),

                    border_color=(
                        0.45,
                        0.90,
                        0.45,
                        1
                    ),

                    size_hint_x=0.16
                )

                # =============================================
                # A BUTTON
                # =============================================

                a = StatusButton(
                    text="A",
                    status_code="A",

                    selected_color=(
                        1.00,
                        0.55,
                        0.55,
                        1
                    ),

                    border_color=(
                        1.00,
                        0.30,
                        0.30,
                        1
                    ),

                    size_hint_x=0.16
                )

                # =============================================
                # H BUTTON
                # =============================================

                h = StatusButton(
                    text="H",
                    status_code="H",

                    selected_color=(
                        0.45,
                        0.65,
                        1.00,
                        1
                    ),

                    border_color=(
                        0.35,
                        0.55,
                        1.00,
                        1
                    ),

                    size_hint_x=0.16
                )

                # =============================================
                # INITIAL SELECTION
                # =============================================

                p.set_selected(
                    existing == "P"
                )

                a.set_selected(
                    existing == "A"
                )

                h.set_selected(
                    existing == "H"
                )

                # =============================================
                # P CLICK
                # =============================================

                def mark_p(
                    instance,
                    label=status,
                    p_button=p,
                    a_button=a,
                    h_button=h
                ):

                    if label.text == "P":

                        # REMOVE P

                        label.text = ""

                        p_button.set_selected(
                            False
                        )

                    else:

                        # SELECT P

                        label.text = "P"

                        label.color = (
                            0.35,
                            1,
                            0.35,
                            1
                        )

                        p_button.set_selected(
                            True
                        )

                        a_button.set_selected(
                            False
                        )

                        h_button.set_selected(
                            False
                        )

                    self.update_summary()

                # =============================================
                # A CLICK
                # =============================================

                def mark_a(
                    instance,
                    label=status,
                    p_button=p,
                    a_button=a,
                    h_button=h
                ):

                    if label.text == "A":

                        # REMOVE A

                        label.text = ""

                        a_button.set_selected(
                            False
                        )

                    else:

                        # SELECT A

                        label.text = "A"

                        label.color = (
                            1,
                            0.35,
                            0.35,
                            1
                        )

                        p_button.set_selected(
                            False
                        )

                        a_button.set_selected(
                            True
                        )

                        h_button.set_selected(
                            False
                        )

                    self.update_summary()

                # =============================================
                # H CLICK
                # =============================================

                def mark_h(
                    instance,
                    label=status,
                    p_button=p,
                    a_button=a,
                    h_button=h
                ):

                    if label.text == "H":

                        # REMOVE H

                        label.text = ""

                        h_button.set_selected(
                            False
                        )

                    else:

                        # SELECT H

                        label.text = "H"

                        label.color = (
                            0.40,
                            0.60,
                            1,
                            1
                        )

                        p_button.set_selected(
                            False
                        )

                        a_button.set_selected(
                            False
                        )

                        h_button.set_selected(
                            True
                        )

                    self.update_summary()

                # =============================================
                # BIND BUTTONS
                # =============================================

                p.bind(
                    on_press=mark_p
                )

                a.bind(
                    on_press=mark_a
                )

                h.bind(
                    on_press=mark_h
                )

                # =============================================
                # ADD TO ROW
                # =============================================

                row_box.add_widget(
                    name_label
                )

                row_box.add_widget(
                    status
                )

                row_box.add_widget(
                    p
                )

                row_box.add_widget(
                    a
                )

                row_box.add_widget(
                    h
                )

                # =============================================
                # RECORD
                # =============================================

                self.records.append({
                    "row": row,
                    "status": status
                })

                self.student_list.add_widget(
                    row_box
                )

                # Staggered student-row entrance animation.
                row_box.opacity = 0
                Clock.schedule_once(
                    lambda dt, widget=row_box:
                    Animation(opacity=1, duration=0.18).start(widget),
                    min(0.45, len(self.records) * 0.015)
                )

            wb.close()

            self.update_summary()

            Animation(opacity=1, duration=0.20).start(self.student_list)

        except Exception as e:
            if hasattr(self, "student_list"):
                Animation(opacity=1, duration=0.10).start(self.student_list)

            self.show_error(
                str(e)
            )

    # ========================================================
    # SUMMARY
    # ========================================================

    def update_summary(self):

        present = 0
        absent = 0
        holiday = 0

        for record in self.records:

            value = record[
                "status"
            ].text

            if value == "P":
                present += 1

            elif value == "A":
                absent += 1

            elif value == "H":
                holiday += 1

        self.present_label.text = (
            f"Present: {present}"
        )

        self.absent_label.text = (
            f"Absent: {absent}"
        )

        self.holiday_label.text = (
            f"Holiday: {holiday}"
        )

    # ========================================================
    # PREVIOUS
    # ========================================================

    def previous_date(self, instance):

        self.load_date(
            self.current_date
            - timedelta(days=1)
        )

    # ========================================================
    # NEXT
    # ========================================================

    def next_date(self, instance):

        self.load_date(
            self.current_date
            + timedelta(days=1)
        )

    # ========================================================
    # TODAY
    # ========================================================

    def go_today(self, instance):

        self.load_date(
            datetime.now().date()
        )

    # ========================================================
    # SAVE
    # ========================================================

    def save_current(self, instance):

        try:

            column = self.find_date_column(
                self.current_date
            )

            if column is None:

                self.show_error(
                    "Selected date was not found in Excel."
                )

                return

            wb = load_workbook(
                self.file_path
            )

            ws = wb["Sheet2"]

            saved = 0

            for record in self.records:

                value = record[
                    "status"
                ].text

                # ---------------------------------------------
                # SELECTED STATUS
                # ---------------------------------------------

                if value in [
                    "P",
                    "A",
                    "H"
                ]:

                    ws.cell(
                        record["row"],
                        column
                    ).value = value

                # ---------------------------------------------
                # DESELECTED = BLANK
                # ---------------------------------------------

                else:

                    ws.cell(
                        record["row"],
                        column
                    ).value = None

                saved += 1

            wb.save(
                self.file_path
            )

            wb.close()

            popup = Popup(
                title="Saved",

                content=Label(
                    text=(
                        self.current_date.strftime(
                            "%d %B %Y"
                        )
                        + "\n\nAttendance Saved!"
                        + f"\n\n{saved} students updated."
                    ),
                    font_size=38
                ),

                size_hint=(
                    0.85,
                    0.35
                )
            )

            popup.open()
            if popup.content:
                popup.content.opacity = 0
                Clock.schedule_once(
                    lambda dt: Animation(opacity=1, duration=0.25).start(popup.content),
                    0
                )

        except Exception as e:

            self.show_error(
                str(e)
            )

    # ========================================================
    # REPORT
    # ========================================================

    def show_report(self, instance):

        try:

            wb = load_workbook(
                self.file_path,
                data_only=True
            )

            ws = wb["Sheet2"]

            report_box = BoxLayout(
                orientation="vertical",
                spacing=8,
                padding=8
            )

            # ------------------------------------------------
            # TITLE
            # ------------------------------------------------

            report_box.add_widget(
                Label(
                    text="STUDENT ATTENDANCE REPORT",
                    font_size=43,
                    size_hint_y=None,
                    height=50
                )
            )

            # ------------------------------------------------
            # SCROLL
            # ------------------------------------------------

            scroll = ScrollView(
                do_scroll_x=False
            )

            report_list = BoxLayout(
                orientation="vertical",
                size_hint_y=None,
                spacing=28,
                padding=5
            )

            report_list.bind(
                minimum_height=report_list.setter(
                    "height"
                )
            )

            # ------------------------------------------------
            # STUDENTS
            # ------------------------------------------------

            for row in range(
                5,
                ws.max_row + 1
            ):

                name = ws.cell(
                    row,
                    2
                ).value

                if name is None:
                    continue

                name = str(
                    name
                ).strip()

                if name == "":
                    continue

                present = 0
                absent = 0
                holiday = 0

                # --------------------------------------------
                # ATTENDANCE DATA
                # --------------------------------------------

                for col in range(
                    4,
                    ws.max_column + 1
                ):

                    value = ws.cell(
                        row,
                        col
                    ).value

                    if value is None:
                        continue

                    value = str(
                        value
                    ).strip().upper()

                    if value == "P":
                        present += 1

                    elif value == "A":
                        absent += 1

                    elif value == "H":
                        holiday += 1

                total = (
                    present
                    + absent
                )

                if total > 0:

                    percentage = (
                        present
                        / total
                    ) * 100

                else:

                    percentage = 0

                text = (
                f"[b]{name:<12}[/b] P:{present} A:{absent} H:{holiday}\n" 
                f"{'':<12} Total:{total} Attendance:{percentage:.1f}%")
                report_list.add_widget(
                Label(
                		text=text,
                		markup=True,
                		font_size=35,
                		halign="left",
                		valign="middle",
                		text_size=(None, None),
                		size_hint_y=None,
                		height=55
        
    )
)

            scroll.add_widget(
                report_list
            )

            report_box.add_widget(
                scroll
            )

            # ------------------------------------------------
            # CLOSE
            # ------------------------------------------------

            close = Button(
                text="CLOSE",
                font_size=29,
                size_hint_y=None,
                height=55
            )

            report_box.add_widget(
                close
            )

            popup = Popup(
                title="Attendance Report",
                content=report_box,
                size_hint=(
                    0.95,
                    0.90
                )
            )

            close.bind(
                on_press=popup.dismiss
            )

            popup.open()

            wb.close()

        except Exception as e:

            self.show_error(
                str(e)
            )

    # ========================================================
    # ERROR POPUP
    # ========================================================

    def show_error(self, message):

        popup = Popup(
            title="Error",

            content=Label(
                text=(
                    "Error:\n\n"
                    + str(message)
                ),
                font_size=17
            ),

            size_hint=(
                0.90,
                0.45
            )
        )

        popup.open()


# ============================================================
# RUN APP
# ============================================================

AttendanceApp().run()